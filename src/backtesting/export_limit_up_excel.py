"""Export reconstructed historical limit-up records to an Excel workbook."""

from __future__ import annotations

import argparse
import sqlite3
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill


SHANGHAI = ZoneInfo("Asia/Shanghai")
DETAIL_HEADERS = [
    "日期",
    "股票代码",
    "股票名称",
    "行业",
    "交易所",
    "涨停制度",
    "连板数",
    "是否ST",
    "开盘价",
    "最高价",
    "最低价",
    "收盘价",
    "昨收价",
    "涨跌幅(%)",
    "换手率(%)",
    "成交额(元)",
    "成交量(股)",
    "一字板(日线判断)",
]


def export_limit_up_excel(
    database: str | Path,
    output: str | Path,
    start_date: date,
    end_date: date,
) -> tuple[Path, int]:
    database = Path(database)
    output = Path(output)
    if not database.exists():
        raise FileNotFoundError(database)
    if start_date > end_date:
        raise ValueError("开始日期不能晚于结束日期")
    output.parent.mkdir(parents=True, exist_ok=True)

    connection = sqlite3.connect(database)
    connection.row_factory = sqlite3.Row
    trading_dates = [
        row[0]
        for row in connection.execute(
            """
            SELECT DISTINCT trade_date FROM daily_bars
            WHERE trade_date BETWEEN ? AND ? ORDER BY trade_date
            """,
            (start_date.isoformat(), end_date.isoformat()),
        )
    ]
    if not trading_dates:
        connection.close()
        raise ValueError("指定区间内没有行情数据")
    date_index = {value: index for index, value in enumerate(trading_dates)}

    workbook = Workbook(write_only=True)
    details = workbook.create_sheet("涨停明细")
    details.freeze_panes = "A2"
    details.append(_header_row(details, DETAIL_HEADERS))
    widths = [13, 12, 16, 24, 10, 15, 10, 10, 12, 12, 12, 12, 12, 13, 13, 18, 18, 18]
    for index, width in enumerate(widths, start=1):
        details.column_dimensions[_column_letter(index)].width = width

    daily: dict[str, dict[str, object]] = {}
    annual: dict[str, dict[str, int]] = defaultdict(
        lambda: {"count": 0, "stocks": 0, "st": 0, "first": 0, "continued": 0}
    )
    last_limit_index: dict[str, int] = {}
    board_count_by_code: dict[str, int] = {}
    record_count = 0
    query = connection.execute(
        """
        SELECT b.trade_date, b.code, s.name, s.industry,
            b.open, b.high, b.low, b.close, b.preclose, b.volume, b.amount,
            b.turnover_rate, b.pct_change, b.is_st
        FROM daily_bars b JOIN stocks s ON s.code=b.code
        WHERE b.trade_date BETWEEN ? AND ?
          AND b.trade_status=1
          AND (
            ((b.code LIKE 'sh.688%' OR b.code LIKE 'sz.300%'
                OR b.code LIKE 'sz.301%')
                AND b.pct_change BETWEEN 19.5 AND 21.0)
            OR
            ((b.code NOT LIKE 'sh.688%' AND b.code NOT LIKE 'sz.300%'
                AND b.code NOT LIKE 'sz.301%')
                AND b.is_st=0 AND b.pct_change BETWEEN 9.5 AND 11.0)
            OR
            ((b.code NOT LIKE 'sh.688%' AND b.code NOT LIKE 'sz.300%'
                AND b.code NOT LIKE 'sz.301%')
                AND b.is_st=1 AND b.pct_change BETWEEN 4.5 AND 5.5)
          )
        ORDER BY b.trade_date, b.code
        """,
        (start_date.isoformat(), end_date.isoformat()),
    )
    for row in query:
        current_index = date_index[row["trade_date"]]
        code = row["code"]
        if last_limit_index.get(code) == current_index - 1:
            board_count = board_count_by_code.get(code, 1) + 1
        else:
            board_count = 1
        last_limit_index[code] = current_index
        board_count_by_code[code] = board_count

        growth_board = code.startswith(("sh.688", "sz.300", "sz.301"))
        if growth_board:
            limit_rule = "20%涨停"
        elif row["is_st"]:
            limit_rule = "ST 5%涨停"
        else:
            limit_rule = "10%涨停"
        one_price = (
            row["open"] == row["high"] == row["low"] == row["close"]
        )
        details.append(
            [
                row["trade_date"],
                code.split(".")[-1],
                row["name"],
                row["industry"] or "未分类",
                "上交所" if code.startswith("sh.") else "深交所",
                limit_rule,
                board_count,
                "是" if row["is_st"] else "否",
                row["open"],
                row["high"],
                row["low"],
                row["close"],
                row["preclose"],
                row["pct_change"],
                row["turnover_rate"],
                row["amount"],
                row["volume"],
                "是" if one_price else "否",
            ]
        )
        record_count += 1

        day = daily.setdefault(
            row["trade_date"],
            {
                "count": 0,
                "st": 0,
                "first": 0,
                "continued": 0,
                "industries": defaultdict(int),
            },
        )
        day["count"] += 1
        day["st"] += int(bool(row["is_st"]))
        day["first"] += int(board_count == 1)
        day["continued"] += int(board_count >= 2)
        day["industries"][row["industry"] or "未分类"] += 1

        year = annual[row["trade_date"][:4]]
        year["count"] += 1
        year["st"] += int(bool(row["is_st"]))
        year["first"] += int(board_count == 1)
        year["continued"] += int(board_count >= 2)

    connection.close()
    details.auto_filter.ref = f"A1:R{record_count + 1}"

    daily_sheet = workbook.create_sheet("每日统计")
    daily_headers = ["日期", "涨停总数", "非ST涨停", "ST涨停", "首板数", "连板数", "涨停最多行业TOP3"]
    daily_sheet.append(_header_row(daily_sheet, daily_headers))
    for trade_date, values in sorted(daily.items()):
        top_sectors = sorted(
            values["industries"].items(), key=lambda item: (-item[1], item[0])
        )[:3]
        daily_sheet.append(
            [
                trade_date,
                values["count"],
                values["count"] - values["st"],
                values["st"],
                values["first"],
                values["continued"],
                "；".join(f"{name}({count})" for name, count in top_sectors),
            ]
        )
    daily_sheet.freeze_panes = "A2"
    daily_sheet.auto_filter.ref = f"A1:G{len(daily) + 1}"

    annual_sheet = workbook.create_sheet("年度统计")
    annual_headers = ["年度", "交易日数", "涨停记录数", "日均涨停数", "ST涨停数", "首板数", "连板数"]
    annual_sheet.append(_header_row(annual_sheet, annual_headers))
    for year, values in sorted(annual.items()):
        year_days = sum(day.startswith(year) for day in daily)
        annual_sheet.append(
            [
                year,
                year_days,
                values["count"],
                round(values["count"] / year_days, 2) if year_days else 0,
                values["st"],
                values["first"],
                values["continued"],
            ]
        )

    notes = workbook.create_sheet("字段说明")
    note_rows = [
        ("文件生成时间", datetime.now(SHANGHAI).isoformat(timespec="seconds")),
        ("实际数据区间", f"{trading_dates[0]} 至 {trading_dates[-1]}"),
        ("数据来源", "BaoStock 沪深A股不复权日线"),
        ("记录数量", record_count),
        ("涨停判定", "主板非ST 9.5%-11%；主板ST 4.5%-5.5%；科创板/创业板 19.5%-21%"),
        ("连板数", "同一股票在相邻市场交易日连续满足上述涨停条件的次数"),
        ("一字板", "仅根据日线开盘价=最高价=最低价=收盘价判断"),
        ("重要限制1", "日线重建结果不含封单金额、首次/最终封板时间及精确炸板次数"),
        ("重要限制2", "BaoStock不覆盖北交所；行业为下载时的证监会行业分类"),
        ("重要限制3", "新股无涨跌幅限制期及规则历史变化可能造成少量误判，不能替代交易所逐笔数据"),
        ("用途", "研究与回测，不构成投资建议"),
    ]
    notes.append(_header_row(notes, ["项目", "说明"]))
    for row in note_rows:
        notes.append(row)
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 100

    workbook.save(output)
    return output, record_count


def _header_row(sheet, values: list[str]) -> list[WriteOnlyCell]:
    cells = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")
        cells.append(cell)
    return cells


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="导出近五年涨停板Excel")
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--start", required=True, type=date.fromisoformat)
    parser.add_argument("--end", required=True, type=date.fromisoformat)
    return parser


def main(argv: Iterable[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    output, count = export_limit_up_excel(
        args.database, args.output, args.start, args.end
    )
    print(f"已导出 {count:,} 条涨停记录: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
