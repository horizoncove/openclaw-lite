#!/usr/bin/env python3
"""Import brands from Excel and fetch/configure logos."""
from __future__ import annotations

import argparse
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path("/home/ubuntu/.cursor/projects/workspace/uploads/Xpark_________0b52.xlsx")
LOGOS_DIR = ROOT / "assets" / "logos"
DATA_DIR = ROOT / "data"

DOMAIN_MAP: dict[str, list[str]] = {
    "蜜雪冰城": ["www.mxbc.com", "mxcoffee.com"],
    "瑞幸": ["luckincoffee.com"],
    "麦当劳": ["mcdonalds.com"],
    "肯德基": ["kfc.com"],
    "海底捞（校园店/嗨捞）": ["haidilao.com", "www.haidilao.com"],
    "海伦司": ["helens.com.cn"],
    "好特卖": ["www.hotmaxx.cn", "hotmaxx.cn"],
    "嗨特购": ["hitgoo.com", "www.hitgoo.com"],
    "魅KTV": ["meiktv.com"],
    "罗森": ["lawson.co.jp"],
    "唐久/每一天": ["www.mytian.com.cn", "mytian.com.cn", "tangjiu.com.cn"],
    "魏家凉皮": ["weijiamedia.com"],
    "霸王茶姬": ["chagee.com", "www.chagee.com"],
    "茶百道": ["chabaidao.com", "www.chabaidao.com"],
    "古茗": ["gumingnc.com", "www.gumingnc.com"],
    "沪上阿姨": ["www.hsay.com", "hsay.com", "hushangayi.com"],
    "书亦烧仙草": ["shuyisxc.com", "www.shuyi-tea.com", "shuyi-tea.com"],
    "CoCo都可": ["coco-tea.com", "www.coco-tea.com.tw"],
    "茶话弄": ["chahuonong.com", "www.chahuonong.com"],
    "茉酸奶": ["moreyogurt.com", "moreyogurt.com.cn"],
    "库迪": ["cotticoffee.com"],
    "幸运咖": ["luckincoffee.com", "xingyunka.com"],
    "塔斯汀": ["tastien.com", "tasiting.com"],
    "正新鸡排": ["zhengxin.com", "zhengxinfood.com"],
    "华莱士": ["hlschina.com", "cnhls.com"],
    "米村拌饭": ["micun.com"],
    "乡村基": ["csc100.com"],
    "杨国福": ["yangguofu.com"],
    "张亮": ["zhangliangmalatang.com"],
    "袁记云饺": ["yuanji.com", "www.yuanji.com"],
    "遇见小面": ["meetnoodle.com"],
    "好利来": ["holiland.com"],
    "鲍师傅": ["baoxianfeng.com", "baoshifu.com"],
    "御品轩": ["yupinxuan.com"],
    "马路边边": ["malubianbian.com"],
    "袁记串串香": ["yuanjichuan.com"],
    "九田家": ["jiutianjia.com"],
    "酒拾烤肉": ["jiushikaorou.com"],
    "半天妖": ["bantianyaoguo.com", "bantianyao.com"],
    "兰湘子": ["lanxiangzi.com"],
    "Perry's": ["perrys.com.cn"],
    "乐刻": ["leoao.com"],
    "TOP TOY": ["toptoyglobal.com"],
    "名创优品": ["miniso.com"],
    "甜啦啦": ["tianlala.com"],
    "益禾堂": ["yht.com", "yhtea.com"],
    "柠季/LINLEE": ["linlee.com.cn"],
    "绝味鸭脖": ["juewei.cn", "juewei.com"],
    "周黑鸭": ["zhouheiya.com", "zhouheiya.cn"],
    "夸父炸串": ["kuafuzhachuan.com"],
    "鲜芋仙": ["meetfresh.com", "www.meetfresh.com"],
    "满记甜品": ["honeydew.com.cn"],
    "The Green Party": ["thegreenparty.com.cn"],
    "三福时尚": ["sanfu.com"],
    "怡康/老百姓大药房": ["lbx.com", "lbxdrugs.com"],
    "Manner": ["mannercoffee.com"],
    "M Stand": ["mstand.cn"],
    "喜茶": ["heytea.com"],
    "奈雪": ["naixuecha.com", "www.naixuecha.com"],
    "达美乐": ["dominos.com"],
    "赛百味": ["subway.com"],
    "怂火锅": ["songhotpot.com"],
    "西塔老太太": ["xitalaotaitai.com"],
    "太二酸菜鱼": ["taier.com"],
    "超级猩猩": ["supermonkey.com.cn"],
    "泡泡玛特": ["popmart.com"],
    "小米之家": ["mi.com"],
    "滔搏/胜道体育（多品牌奥莱店）": ["yysports.com", "topsports.com.cn"],
    "李宁奥莱店/工厂店": ["lining.com"],
    "安踏奥莱店": ["anta.com"],
    "特步工厂店": ["xtep.com.cn", "xtep.com"],
    "361°奥莱店": ["361sport.com"],
    "鸿星尔克": ["www.erke.com", "erke.com"],
    "斯凯奇奥莱店": ["skechers.com"],
    "New Balance奥莱店": ["newbalance.com"],
    "Nike Factory Store（直营奥莱）": ["nike.com"],
    "Adidas Factory Outlet": ["adidas.com"],
    "PUMA奥莱店": ["puma.com"],
    "Under Armour奥莱店": ["underarmour.com"],
    "The North Face/始祖鸟折扣": ["thenorthface.com", "arcteryx.com"],
    "匡威/Vans折扣店": ["converse.com", "vans.com"],
    "骆驼/探路者折扣店": ["camel.com.cn", "toread.com.cn"],
    "COMMUNE公社": ["commune.com.cn"],
    "汤姆熊/大玩家电玩城": ["tomworld.com.cn", "dawanjia.com"],
    "一只酸奶牛": ["yizhisuannai.com"],
    "电竞馆/网咖（杰拉/网鱼）": ["www.wywk.cn", "wywk.cn", "wangyu.com"],
    "网鱼网咖": ["wangyu.com"],
}

WIKI_MAP: dict[str, str] = {
    "瑞幸": "Luckin Coffee",
    "麦当劳": "McDonald's",
    "肯德基": "KFC",
    "海底捞（校园店/嗨捞）": "Haidilao",
    "喜茶": "Heytea",
    "奈雪": "Nayuki",
    "达美乐": "Domino's Pizza",
    "赛百味": "Subway (restaurant)",
    "名创优品": "Miniso",
    "罗森": "Lawson (store)",
    "小米之家": "Xiaomi",
    "泡泡玛特": "Pop Mart",
    "李宁奥莱店/工厂店": "Li-Ning",
    "安踏奥莱店": "Anta Sports",
    "Nike Factory Store（直营奥莱）": "Nike, Inc.",
    "Adidas Factory Outlet": "Adidas",
    "PUMA奥莱店": "Puma (brand)",
    "New Balance奥莱店": "New Balance",
    "Under Armour奥莱店": "Under Armour",
    "The North Face/始祖鸟折扣": "The North Face",
    "匡威/Vans折扣店": "Converse (shoe company)",
    "斯凯奇奥莱店": "Skechers",
    "鸿星尔克": "Erke (brand)",
    "特步工厂店": "Xtep",
    "361°奥莱店": "361 Degrees",
    "满记甜品": "Honeymoon Dessert",
    "鲜芋仙": "Meet Fresh",
    "周黑鸭": "Zhou Hei Ya",
    "绝味鸭脖": "Juewei Food",
    "华莱士": "Wallace (fast food)",
}

TIER_TAG = {"A": "coral", "B": "lime", "C": "ink-outline", "D": "ink-outline"}
PROB_TAG = {"高": "lime", "中高": "lime", "中": "ink-outline", "低": "ink-outline"}


def slugify(name: str, idx: int) -> str:
    base = re.sub(r"[^\w\u4e00-\u9fff]+", "-", name.strip()).strip("-").lower()
    if not base or len(base) > 40:
        base = f"brand-{idx:03d}"
    return base[:48]


def fetch_url(url: str, timeout: int = 10) -> bytes | None:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 XparkLogoBot/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
            if len(data) < 200:
                return None
            return data
    except (urllib.error.URLError, TimeoutError, OSError):
        return None


def try_domain_sources(domain: str) -> tuple[bytes | None, str]:
    candidates = [
        f"https://{domain}/favicon.ico",
        f"https://www.{domain}/favicon.ico",
        (
            "https://t0.gstatic.com/faviconV2?client=SOCIAL&type=FAVICON"
            f"&fallback_opts=TYPE,SIZE,URL&url=https://{domain}&size=256"
        ),
        f"https://icons.duckduckgo.com/ip3/{domain}.ico",
        f"https://www.google.com/s2/favicons?domain={domain}&sz=256",
        f"https://api.faviconkit.com/{domain}/256",
        f"https://logo.clearbit.com/{domain}",
    ]
    for url in candidates:
        data = fetch_url(url)
        if data and len(data) >= 400:
            return data, "png"
    return None, "png"


def try_wikipedia(name: str) -> tuple[bytes | None, str]:
    title = WIKI_MAP.get(name)
    if not title:
        return None, "png"
    api = (
        "https://en.wikipedia.org/w/api.php?action=query&titles="
        + urllib.parse.quote(title)
        + "&prop=pageimages&format=json&pithumbsize=256"
    )
    raw = fetch_url(api)
    if not raw:
        return None, "png"
    try:
        payload = json.loads(raw)
        pages = payload["query"]["pages"]
        for page in pages.values():
            thumb = page.get("thumbnail")
            if thumb and thumb.get("source"):
                img = fetch_url(thumb["source"])
                if img and len(img) >= 400:
                    return img, "png"
    except (KeyError, json.JSONDecodeError):
        pass
    return None, "png"


def fetch_brand_logo(name: str) -> tuple[bytes | None, str]:
    for domain in DOMAIN_MAP.get(name, []):
        data, ext = try_domain_sources(domain)
        if data:
            return data, ext
    return try_wikipedia(name)


def make_svg_placeholder(name: str, tier: str) -> str:
    short = re.sub(r"[（(/].*", "", name).strip()
    label = short[:4] if len(short) > 4 else short
    colors = {"A": "#FF5436", "B": "#A0E828", "C": "#1A1918", "D": "#6B7280"}
    fill = colors.get(tier, "#1A1918")
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 160 160" role="img">
  <rect width="160" height="160" rx="20" fill="#FEFCF8"/>
  <rect x="12" y="12" width="136" height="136" rx="16" fill="{fill}" opacity="0.1"/>
  <text x="80" y="92" text-anchor="middle" font-family="'Noto Sans SC',sans-serif" font-size="28" font-weight="700" fill="{fill}">{label}</text>
</svg>'''


def resolve_logo(name: str, tier: str, slug: str, fill_missing: bool = False) -> str:
    png_path = LOGOS_DIR / f"{slug}.png"
    svg_path = LOGOS_DIR / f"{slug}.svg"

    if fill_missing and png_path.exists():
        return f"logos/{slug}.png"
    if fill_missing and svg_path.exists():
        existing = fetch_brand_logo(name)
        if existing[0]:
            png_path.write_bytes(existing[0])
            svg_path.unlink(missing_ok=True)
            return f"logos/{slug}.png"

    data, ext = fetch_brand_logo(name)
    if data:
        path = LOGOS_DIR / f"{slug}.{ext}"
        path.write_bytes(data)
        svg_path.unlink(missing_ok=True)
        return f"logos/{slug}.{ext}"

    if fill_missing and svg_path.exists():
        return f"logos/{slug}.svg"

    svg = make_svg_placeholder(name, tier)
    svg_path.write_text(svg, encoding="utf-8")
    png_path.unlink(missing_ok=True)
    return f"logos/{slug}.svg"


def load_brands(fill_missing: bool = False) -> list[dict]:
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["品牌总清单"]
    brands = []
    idx = 1
    for row in ws.iter_rows(min_row=5, values_only=True):
        tier, cat, subcat, name, prob, prob_desc, audience, floor, area, rent = row[1:11]
        if not name:
            continue
        name = str(name).strip()
        slug = slugify(name, idx)
        logo = resolve_logo(name, str(tier), slug, fill_missing=fill_missing)
        brands.append({
            "id": f"{idx:03d}",
            "slug": slug,
            "name": name,
            "tier": str(tier),
            "cat": str(cat),
            "subcat": str(subcat or ""),
            "prob": str(prob or ""),
            "probDesc": str(prob_desc or ""),
            "audience": str(audience or ""),
            "floor": str(floor or ""),
            "area": str(area or ""),
            "rent": str(rent or ""),
            "logo": logo,
            "tierTag": TIER_TAG.get(str(tier), "ink-outline"),
            "probTag": PROB_TAG.get(str(prob), "ink-outline"),
        })
        idx += 1
    return brands


def write_outputs(brands: list[dict]) -> None:
    LOGOS_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    (DATA_DIR / "brands.json").write_text(
        json.dumps(brands, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    js = "window.XPARK_BRANDS = " + json.dumps(brands, ensure_ascii=False, indent=2) + ";\n"
    (ROOT / "js" / "brands-data.js").write_text(js, encoding="utf-8")

    png = sum(1 for b in brands if b["logo"].endswith(".png"))
    svg = sum(1 for b in brands if b["logo"].endswith(".svg"))
    cats = sorted({b["cat"] for b in brands})
    meta = {"total": len(brands), "pngLogos": png, "svgLogos": svg, "categories": cats}
    (DATA_DIR / "brands-meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Exported {len(brands)} brands — PNG logos: {png}, SVG placeholders: {svg}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--fill-missing", action="store_true", help="Only fetch logos for SVG placeholders")
    args = parser.parse_args()
    brands = load_brands(fill_missing=args.fill_missing)
    write_outputs(brands)
