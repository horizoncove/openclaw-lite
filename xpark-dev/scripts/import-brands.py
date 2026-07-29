#!/usr/bin/env python3
"""Import brands from Excel and fetch/configure logos."""
from __future__ import annotations

import json
import re
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL = Path("/home/ubuntu/.cursor/projects/workspace/uploads/Xpark_________0b52.xlsx")
LOGOS_DIR = ROOT / "assets" / "logos"
DATA_DIR = ROOT / "data"

# Brand name -> domain for logo lookup (Clearbit / favicon)
# Brand name -> candidate domains (first match wins)
DOMAIN_MAP: dict[str, list[str]] = {
    "蜜雪冰城": ["www.mxbc.com", "mxcoffee.com"],
    "瑞幸": ["luckincoffee.com"],
    "麦当劳": ["mcdonalds.com"],
    "肯德基": ["kfc.com"],
    "海底捞（校园店/嗨捞）": ["haidilao.com", "www.haidilao.com"],
    "海伦司": ["helens.com.cn"],
    "好特卖": ["hotmaxx.cn"],
    "嗨特购": ["hitgoo.com"],
    "魅KTV": ["meiktv.com"],
    "罗森": ["lawson.co.jp"],
    "唐久/每一天": ["tangjiu.com.cn"],
    "魏家凉皮": ["weijiamedia.com"],
    "霸王茶姬": ["chagee.com", "www.chagee.cn"],
    "茶百道": ["chabaidao.com", "www.chabaidao.com"],
    "古茗": ["gumingnc.com"],
    "沪上阿姨": ["hushangayi.com"],
    "书亦烧仙草": ["shuyi-tea.com"],
    "CoCo都可": ["coco-tea.com.tw"],
    "茶话弄": ["chahuonong.com"],
    "茉酸奶": ["moreyogurt.com.cn"],
    "库迪": ["cotticoffee.com"],
    "幸运咖": ["xingyunka.com"],
    "塔斯汀": ["tasiting.com", "tastien.com"],
    "正新鸡排": ["zhengxinfood.com"],
    "华莱士": ["hlschina.com"],
    "米村拌饭": ["micun.com"],
    "乡村基": ["csc100.com"],
    "杨国福": ["yangguofu.com"],
    "张亮": ["zhangliangmalatang.com"],
    "袁记云饺": ["yuanji.com"],
    "遇见小面": ["meetnoodle.com"],
    "好利来": ["holiland.com"],
    "鲍师傅": ["baoshifu.com"],
    "马路边边": ["malubianbian.com"],
    "九田家": ["jiutianjia.com"],
    "半天妖": ["bantianyao.com"],
    "兰湘子": ["lanxiangzi.com"],
    "Perry's": ["perrys.com.cn"],
    "乐刻": ["leoao.com"],
    "TOP TOY": ["toptoyglobal.com"],
    "名创优品": ["miniso.com"],
    "甜啦啦": ["tianlala.com"],
    "益禾堂": ["yht.com"],
    "柠季/LINLEE": ["linlee.com.cn"],
    "绝味鸭脖": ["juewei.cn"],
    "周黑鸭": ["zhouheiya.cn"],
    "夸父炸串": ["kuafuzhachuan.com"],
    "鲜芋仙": ["meetfresh.com"],
    "满记甜品": ["honeydew.com.cn"],
    "The Green Party": ["thegreenparty.com.cn"],
    "三福时尚": ["sanfu.com"],
    "怡康/老百姓大药房": ["lbxdrugs.com"],
    "Manner": ["mannercoffee.com"],
    "M Stand": ["mstand.cn"],
    "喜茶": ["heytea.com"],
    "奈雪": ["naixuecha.com"],
    "达美乐": ["dominos.com"],
    "赛百味": ["subway.com"],
    "怂火锅": ["songhotpot.com"],
    "西塔老太太": ["xitalaotaitai.com"],
    "太二酸菜鱼": ["taier.com"],
    "超级猩猩": ["supermonkey.com.cn"],
    "泡泡玛特": ["popmart.com"],
    "小米之家": ["mi.com"],
    "滔搏/胜道体育（多品牌奥莱店）": ["topsports.com.cn"],
    "李宁奥莱店/工厂店": ["lining.com"],
    "安踏奥莱店": ["anta.com"],
    "特步工厂店": ["xtep.com.cn"],
    "361°奥莱店": ["361sport.com"],
    "鸿星尔克": ["erke.com"],
    "斯凯奇奥莱店": ["skechers.com"],
    "New Balance奥莱店": ["newbalance.com"],
    "Nike Factory Store（直营奥莱）": ["nike.com"],
    "Adidas Factory Outlet": ["adidas.com"],
    "PUMA奥莱店": ["puma.com"],
    "Under Armour奥莱店": ["underarmour.com"],
    "The North Face/始祖鸟折扣": ["thenorthface.com"],
    "匡威/Vans折扣店": ["converse.com"],
    "骆驼/探路者折扣店": ["camel.com.cn"],
    "COMMUNE公社": ["commune.com.cn"],
    "汤姆熊/大玩家电玩城": ["tomworld.com.cn"],
}

CAT_PHOTO = {
    "茶饮咖啡": "brands/roast.jpg",
    "快餐轻食": "brands/grain.jpg",
    "正餐聚餐": "brands/nori.jpg",
    "夜间经济": "scenes/district-night.jpg",
    "零售": "brands/retail.jpg",
    "文体娱": "events/cinema.jpg",
    "钩子业态": "events/music-market.jpg",
    "运动奥莱": "brands/lifestyle.jpg",
    "生活配套": "brands/retail.jpg",
    "教育培训": "features/curating.jpg",
}

TIER_TAG = {"A": "coral", "B": "lime", "C": "ink-outline", "D": "ink-outline"}
PROB_TAG = {"高": "lime", "中高": "lime", "中": "ink-outline", "低": "ink-outline"}


def slugify(name: str, idx: int) -> str:
    base = re.sub(r"[^\w\u4e00-\u9fff]+", "-", name.strip()).strip("-").lower()
    if not base or len(base) > 40:
        base = f"brand-{idx:03d}"
    return base[:48]


def fetch_url(url: str, timeout: int = 8) -> bytes | None:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 XparkLogoBot/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
            if len(data) < 200:
                return None
            return data
    except (urllib.error.URLError, TimeoutError, OSError):
        return None


def try_fetch_logo(domain: str) -> tuple[bytes | None, str]:
    sources = [
        (f"https://icons.duckduckgo.com/ip3/{domain}.ico", "png"),
        (f"https://www.google.com/s2/favicons?domain={domain}&sz=128", "png"),
        (f"https://logo.clearbit.com/{domain}", "png"),
    ]
    for url, ext in sources:
        data = fetch_url(url)
        if data and len(data) >= 400:
            return data, ext
    return None, "png"


def fetch_brand_logo(name: str) -> tuple[bytes | None, str]:
    domains = DOMAIN_MAP.get(name, [])
    for domain in domains:
        data, ext = try_fetch_logo(domain)
        if data:
            return data, ext
    return None, "png"


def make_svg_placeholder(name: str, tier: str) -> str:
    short = re.sub(r"[（(].*", "", name).strip()
    label = short[:4] if len(short) > 4 else short
    colors = {"A": "#FF5436", "B": "#A0E828", "C": "#1A1918", "D": "#6B7280"}
    fill = colors.get(tier, "#1A1918")
    return f'''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 120 120" role="img">
  <rect width="120" height="120" rx="16" fill="#FEFCF8"/>
  <rect x="8" y="8" width="104" height="104" rx="12" fill="{fill}" opacity="0.12"/>
  <text x="60" y="68" text-anchor="middle" font-family="'Noto Sans SC',sans-serif" font-size="22" font-weight="700" fill="{fill}">{label}</text>
</svg>'''


def resolve_logo(name: str, tier: str, slug: str) -> str:
    filename_base = slug
    data, ext = fetch_brand_logo(name)
    if data:
        path = LOGOS_DIR / f"{filename_base}.{ext}"
        path.write_bytes(data)
        return f"logos/{filename_base}.{ext}"

    svg = make_svg_placeholder(name, tier)
    path = LOGOS_DIR / f"{filename_base}.svg"
    path.write_text(svg, encoding="utf-8")
    return f"logos/{filename_base}.svg"


def load_brands() -> list[dict]:
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["品牌总清单"]
    brands = []
    idx = 1
    for row in ws.iter_rows(min_row=5, values_only=True):
        tier, cat, subcat, name, prob, prob_desc, audience, floor, area, rent = row[1:11]
        if not name:
            continue
        name = str(name).strip()
        slug = slugify(name, idx)
        logo = resolve_logo(name, str(tier), slug)
        brands.append({
            "id": f"{idx:03d}",
            "slug": slug,
            "name": name,
            "tier": str(tier),
            "cat": str(cat),
            "subcat": str(subcat or ""),
            "prob": str(prob or ""),
            "probDesc": str(prob_desc or ""),
            "audience": str(audience or ""),
            "floor": str(floor or ""),
            "area": str(area or ""),
            "rent": str(rent or ""),
            "logo": logo,
            "photo": CAT_PHOTO.get(str(cat), "brands/retail.jpg"),
            "tierTag": TIER_TAG.get(str(tier), "ink-outline"),
            "probTag": PROB_TAG.get(str(prob), "ink-outline"),
        })
        idx += 1
    return brands


def write_outputs(brands: list[dict]) -> None:
    LOGOS_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    (DATA_DIR / "brands.json").write_text(
        json.dumps(brands, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    js = "window.XPARK_BRANDS = " + json.dumps(brands, ensure_ascii=False, indent=2) + ";\n"
    (ROOT / "js" / "brands-data.js").write_text(js, encoding="utf-8")

    cats = sorted({b["cat"] for b in brands})
    tiers = sorted({b["tier"] for b in brands})
    meta = {"total": len(brands), "categories": cats, "tiers": tiers}
    (DATA_DIR / "brands-meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Exported {len(brands)} brands, {len(cats)} categories")
    print("Categories:", cats)


if __name__ == "__main__":
    brands = load_brands()
    write_outputs(brands)
